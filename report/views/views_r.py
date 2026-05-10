import hashlib, uuid, os, datetime, csv
from django.http import Http404
from django.utils import timezone
from django.views.generic import DetailView
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.models import User,Group
from django.contrib.auth.decorators import login_required
from custom.models import  Position, Country, Municipality, AdministrativePost, Status, Village, SubVillage, EducationLevel,\
                           University, Faculty, StudyProgram, Year, Estructure
from django.template.loader import render_to_string
from django.contrib import messages
from membro.models import Membru, LocationTL, ContactInfo, AddressOrigin, MembroPosition
from django.db.models import Count, Sum, Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User,Group
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from config.decorators import allowed_users
from config.auth_utils import c_user_mem
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

@login_required
@allowed_users(allowed_roles=['postu', 'suku', 'ald', 'admin', 'staff'])
def dash(request):
    group = request.user.groups.all()[0].name
    user_mem = c_user_mem(request.user)

    loc = None
    if user_mem:
        loc = LocationTL.objects.select_related(
            'village', 'aldeia', 'municipality'
        ).filter(membro=user_mem).first()

    nama_lokasi = "LOKASAUN"
    scope       = None
    data        = []

    if group in ['postu', 'admin', 'staff']:
        data        = list(Village.objects.all().order_by('name'))
        scope       = 'village'
        nama_lokasi = "POSTU ADMINISTRATIVU"
    elif group == 'suku':
        if loc and loc.village:
            data        = list(SubVillage.objects.filter(village=loc.village).order_by('name'))
            scope       = 'aldeia'
            nama_lokasi = f"SUCO {loc.village.name.upper()}"
    elif group == 'ald':
        if loc and loc.aldeia:
            data        = [loc.aldeia]
            scope       = 'aldeia'
            nama_lokasi = f"ALDEIA {loc.aldeia.name.upper()}"

    def count_pos(qs, keyword):
        return qs.filter(
            membro__membroposition__position__name__icontains=keyword
        ).count()

    # ── Build semua data (untuk export & paginator) ──
    all_objects = []
    total_all = total_mane = total_feto = 0
    total_ativu = total_la_ativu = 0
    total_kord = total_sek = total_mem = 0
    total_ojd = total_mpd = total_opd = 0

    for d in data:
        if scope == 'village':
            qs = LocationTL.objects.filter(village=d)
            label_tipe = "Suku"
        else:
            qs = LocationTL.objects.filter(aldeia=d)
            label_tipe = "Aldeia"

        male     = qs.filter(membro__sex="Mane").count()
        female   = qs.filter(membro__sex="Feto").count()
        total    = male + female
        ativu    = qs.filter(membro__is_appr=True).count()
        la_ativu = qs.filter(membro__is_appr=False).count()
        kord     = count_pos(qs, "Cordenador")
        sek      = count_pos(qs, "Sekretariu")
        mem      = count_pos(qs, "Membro")
        ojd      = count_pos(qs, "OJD")
        mpd      = count_pos(qs, "MPD")
        opd      = count_pos(qs, "OPD")

        all_objects.append({
            "id": d.id, "nama": d.name, "tipe": label_tipe,
            "male": male, "female": female, "total": total,
            "ativu": ativu, "la_ativu": la_ativu,
            "kord": kord, "sek": sek, "mem": mem,
            "ojd": ojd, "mpd": mpd, "opd": opd,
        })

        total_all      += total;  total_mane     += male
        total_feto     += female; total_ativu    += ativu
        total_la_ativu += la_ativu
        total_kord     += kord;   total_sek      += sek
        total_mem      += mem;    total_ojd      += ojd
        total_mpd      += mpd;    total_opd      += opd

    totals = {
        "total_all": total_all, "total_mane": total_mane,
        "total_feto": total_feto, "total_ativu": total_ativu,
        "total_la_ativu": total_la_ativu, "total_kord": total_kord,
        "total_sek": total_sek, "total_mem": total_mem,
        "total_ojd": total_ojd, "total_mpd": total_mpd,
        "total_opd": total_opd,
    }

    # ── Export handler ──
    export = request.GET.get('export')
    export_scope = request.GET.get('scope', 'all')  # 'all' atau 'page'
    page_number  = request.GET.get('page', 1)

    if export in ['excel', 'pdf']:
        if export_scope == 'page':
            paginator   = Paginator(all_objects, 10)
            page_data   = paginator.get_page(page_number)
            export_data = list(page_data)
        else:
            export_data = all_objects

        if export == 'excel':
            return export_excel(export_data, totals, nama_lokasi)
        elif export == 'pdf':
            return export_pdf(export_data, totals, nama_lokasi)

    # ── Paginator untuk tampilan ──
    paginator  = Paginator(all_objects, 10)
    page_obj   = paginator.get_page(page_number)

    context = {
        "title": "Sumario Membru", "legend": "Sumario Membru",
        "group": group, "nama_lokasi": nama_lokasi,
        "objects1": page_obj,       # pakai page_obj di template
        "page_obj": page_obj,
        **totals,
    }
    return render(request, 'Sum/dasht.html', context)


# ── Export Excel ──
def export_excel(data, totals, nama_lokasi):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sumario Membru"

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    total_font  = Font(bold=True, size=9)
    center      = Alignment(horizontal="center", vertical="center")

    # Judul
    ws.merge_cells("A1:N1")
    ws["A1"] = f"SUMARIO MEMBRU — {nama_lokasi}"
    ws["A1"].font = Font(bold=True, size=11)
    ws["A1"].alignment = center

    # Header baris 1
    headers1 = ["#", "Naran", "Mane", "Feto", "Ativu", "La Ativu",
                 "Kord.", "Sek.", "Mem.", "OJD", "MPD", "OPD"]
    for col, h in enumerate(headers1, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Data
    for i, obj in enumerate(data, 1):
        row = [i, obj['nama'], obj['male'], obj['female'],
               obj['ativu'], obj['la_ativu'],
               obj['kord'], obj['sek'], obj['mem'],
               obj['ojd'], obj['mpd'], obj['opd']]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+2, column=col, value=val)
            cell.alignment = center

    # Total row
    total_row = len(data) + 3
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.merge_cells(f"A{total_row}:B{total_row}")
    for col, key in enumerate(['total_mane','total_feto','total_ativu','total_la_ativu',
                                'total_kord','total_sek','total_mem',
                                'total_ojd','total_mpd','total_opd'], 3):
        cell = ws.cell(row=total_row, column=col, value=totals[key])
        cell.font = total_font
        cell.alignment = center

    # Auto width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 8)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sumario_membru.xlsx"'
    wb.save(response)
    return response


# ── Export PDF ──
def export_pdf(data, totals, nama_lokasi):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sumario_membru.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            leftMargin=20, rightMargin=20,
                            topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = []

    # Judul
    elements.append(Paragraph(f"SUMARIO MEMBRU — {nama_lokasi}", styles['Title']))
    elements.append(Spacer(1, 10))

    # Header tabel
    headers = ["#", "Naran", "Mane", "Feto", "Ativu", "La Ativu",
               "Kord.", "Sek.", "Mem.", "OJD", "MPD", "OPD"]
    table_data = [headers]

    for i, obj in enumerate(data, 1):
        table_data.append([
            i, obj['nama'], obj['male'], obj['female'],
            obj['ativu'], obj['la_ativu'],
            obj['kord'], obj['sek'], obj['mem'],
            obj['ojd'], obj['mpd'], obj['opd'],
        ])

    # Total row
    table_data.append([
        "TOTAL", "",
        totals['total_mane'], totals['total_feto'],
        totals['total_ativu'], totals['total_la_ativu'],
        totals['total_kord'], totals['total_sek'], totals['total_mem'],
        totals['total_ojd'], totals['total_mpd'], totals['total_opd'],
    ])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
        ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 8),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('ALIGN',       (1,0), (1,-1),  'LEFT'),
        ('GRID',        (0,0), (-1,-1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f1f5f9')]),
        ('BACKGROUND',  (0,-1), (-1,-1), colors.HexColor('#dbeafe')),
        ('FONTNAME',    (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('SPAN',        (0,-1), (1,-1)),
    ]))
    elements.append(t)
    doc.build(elements)
    return response

@login_required
@allowed_users(allowed_roles=['postu', 'suku', 'ald', 'admin', 'staff'])
def lista_membro(request, sex, loc_id):
    group    = request.user.groups.all()[0].name
    user_mem = c_user_mem(request.user)

    # Tentukan scope berdasarkan group
    if group in ['postu', 'admin', 'staff']:
        qs    = LocationTL.objects.filter(village_id=loc_id)
        nama  = Village.objects.filter(id=loc_id).first()
        label = f"SUKU — {nama}"
    else:
        qs    = LocationTL.objects.filter(aldeia_id=loc_id)
        nama  = SubVillage.objects.filter(id=loc_id).first()
        label = f"ALDEIA — {nama}"

    # Filter sex
    if sex == 'mane':
        qs        = qs.filter(membro__sex="Mane")
        label_sex = "Mane"
    elif sex == 'feto':
        qs        = qs.filter(membro__sex="Feto")
        label_sex = "Feto"
    elif sex == 'la_ativu':
        qs        = qs.filter(membro__is_appr=False)
        label_sex = "La Ativu"
    elif sex == 'ativu':
        qs        = qs.filter(membro__is_appr=True)
        label_sex = "Ativu"
    else:
        label_sex = "Hotu"

    membros = Membru.objects.filter(
        locationtl__in=qs
    ).select_related(
        'contactinfo', 'locationtl__village',
        'locationtl__aldeia', 'membroposition__position', 'status'
    ).order_by('name')

    # ── Export ──
    export = request.GET.get('export')
    if export == 'excel':
        return export_lista_excel(membros, label, label_sex)
    elif export == 'pdf':
        return export_lista_pdf(membros, label, label_sex)

    context = {
        'membros':   membros,
        'group':     group,
        'sex':       label_sex,
        'nama':      nama,
        'estrutura': label,
        'title':     f'Lista Membru — {label}',
        'legend':    f'Lista Membru {label_sex} — {label}',
        'total':     membros.count(),
        'total_mane': membros.filter(sex="Mane").count(),
        'total_feto': membros.filter(sex="Feto").count(),
    }
    return render(request, 'Sum/ista_membro.html', context)


# ── Export Excel ──
# ── Export Excel ── GANTI bagian auto width di bawah
def export_lista_excel(membros, label, label_sex):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter  # ← import ini

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Lista Membro"

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left", vertical="center")

    # Judul
    ws.merge_cells("A1:I1")
    ws["A1"] = f"LISTA MEMBRO {label_sex.upper()} — {label.upper()}"
    ws["A1"].font      = Font(bold=True, size=11)
    ws["A1"].alignment = center

    # Sub judul total
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Total: {membros.count()} Membro"
    ws["A2"].font      = Font(size=9, italic=True)
    ws["A2"].alignment = center

    # Header
    headers = ["#", "Naran", "Sexu", "Data Moris", "Telefone",
               "Suku", "Aldeia", "Pozisaun", "Status"]
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center

    # Data
    for i, m in enumerate(membros, 1):
        phone    = m.contactinfo.phone if hasattr(m, 'contactinfo') and m.contactinfo else "—"
        village  = m.locationtl.village.name if hasattr(m, 'locationtl') and m.locationtl and m.locationtl.village else "—"
        aldeia   = m.locationtl.aldeia.name  if hasattr(m, 'locationtl') and m.locationtl and m.locationtl.aldeia  else "—"
        position = m.membroposition.position.name if hasattr(m, 'membroposition') and m.membroposition and m.membroposition.position else "—"
        status   = str(m.status) if m.status else "—"
        dob      = m.dob.strftime("%d/%m/%Y") if m.dob else "—"

        row_data = [i, m.name, m.sex, dob, phone, village, aldeia, position, status]
        for col, val in enumerate(row_data, 1):
            cell           = ws.cell(row=i+3, column=col, value=val)
            cell.alignment = left if col == 2 else center
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F1F5F9")

    # ✅ GANTI auto width — pakai get_column_letter, bukan ws.columns
    col_widths = [5, 30, 8, 14, 14, 18, 18, 20, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="lista_membro_{label_sex}.xlsx"'
    wb.save(response)
    return response


# ── Export PDF ──
def export_lista_pdf(membros, label, label_sex):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lista_membro_{label_sex}.pdf"'

    doc      = SimpleDocTemplate(response, pagesize=landscape(A4),
                                 leftMargin=20, rightMargin=20,
                                 topMargin=20, bottomMargin=20)
    styles   = getSampleStyleSheet()
    elements = []

    # Judul
    elements.append(Paragraph(
        f"LISTA MEMBRO {label_sex.upper()} — {label.upper()}", styles['Title']
    ))
    elements.append(Paragraph(
        f"Total: {membros.count()} Membro", styles['Normal']
    ))
    elements.append(Spacer(1, 10))

    # Header
    headers = ["#", "Naran", "Sexu", "Data Moris", "Telefone",
               "Suku", "Aldeia", "Pozisaun", "Status"]
    table_data = [headers]

    for i, m in enumerate(membros, 1):
        phone    = m.contactinfo.phone if hasattr(m, 'contactinfo') and m.contactinfo else "—"
        village  = m.locationtl.village.name if hasattr(m, 'locationtl') and m.locationtl and m.locationtl.village else "—"
        aldeia   = m.locationtl.aldeia.name  if hasattr(m, 'locationtl') and m.locationtl and m.locationtl.aldeia  else "—"
        position = m.membroposition.position.name if hasattr(m, 'membroposition') and m.membroposition and m.membroposition.position else "—"
        status   = str(m.status) if m.status else "—"
        dob      = m.dob.strftime("%d/%m/%Y") if m.dob else "—"

        table_data.append([
            i, m.name, m.sex, dob, phone,
            village, aldeia, position, status
        ])

    col_widths = [20, 120, 35, 55, 60, 70, 70, 80, 50]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',     (0,0),  (-1,0),  colors.HexColor('#1A3A5C')),
        ('TEXTCOLOR',      (0,0),  (-1,0),  colors.white),
        ('FONTNAME',       (0,0),  (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0,0),  (-1,-1), 7),
        ('ALIGN',          (0,0),  (-1,-1), 'CENTER'),
        ('ALIGN',          (1,1),  (1,-1),  'LEFT'),
        ('GRID',           (0,0),  (-1,-1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0,1),  (-1,-1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('TOPPADDING',     (0,0),  (-1,-1), 3),
        ('BOTTOMPADDING',  (0,0),  (-1,-1), 3),
    ]))
    elements.append(t)
    doc.build(elements)
    return response

def get_role_queryset(group, loc):
    qs = LocationTL.objects.select_related(
        'village', 'aldeia', 'membro',
        'membro__membroposition__position',
        'membro__status'
    )
    if group in ['postu', 'admin', 'staff']:
        return qs, 'geral'
    elif group == 'suku':
        if loc and loc.village:
            return qs.filter(village=loc.village), 'suku'
        return qs.none(), 'suku'
    elif group == 'ald':
        if loc and loc.aldeia:
            return qs.filter(aldeia=loc.aldeia), 'ald'
        return qs.none(), 'ald'
    return qs.none(), 'none'

@login_required
@allowed_users(allowed_roles=['postu', 'suku', 'ald', 'admin', 'staff'])
def report_menu(request):
    group = request.user.groups.all()[0].name
    context = {
        'title': 'Relatoriu',
        'legend': 'Relatoriu',
        'group': group,
    }
    return render(request, 'Report/menu.html', context)

@login_required
@allowed_users(allowed_roles=['postu', 'suku', 'ald', 'admin', 'staff'])
def report_geral(request):
    group = request.user.groups.all()[0].name
    user_mem = c_user_mem(request.user)
    loc = None
    if user_mem:
        loc = LocationTL.objects.select_related('village', 'aldeia').filter(membro=user_mem).first()
    qs, tipe = get_role_queryset(group, loc)
    if group in ['postu', 'admin', 'staff']:
        groups_data = Village.objects.all()
        def get_qs(d): return qs.filter(village=d)
        label = 'Suku'
    elif group == 'suku':
        groups_data = SubVillage.objects.filter(village=loc.village) if loc and loc.village else []
        def get_qs(d): return qs.filter(aldeia=d)
        label = 'Aldeia'
    elif group == 'ald':
        groups_data = [loc.aldeia] if loc and loc.aldeia else []
        def get_qs(d): return qs.filter(aldeia=d)
        label = 'Aldeia'
    rows = []
    total_mane = total_feto = total_all = 0
    for d in groups_data:
        q = get_qs(d)
        mane   = q.filter(membro__sex='Mane').count()
        feto   = q.filter(membro__sex='Feto').count()
        total  = mane + feto
        rows.append({'nama': d.name, 'mane': mane, 'feto': feto, 'total': total})
        total_mane += mane
        total_feto += feto
        total_all  += total

    context = {
        'title': 'Relatoriu Geral',
        'legend': 'Relatoriu Geral Membro',
        'group': group, 'label': label,
        'rows': rows,
        'total_mane': total_mane,
        'total_feto': total_feto,
        'total_all': total_all,
    }
    return render(request, 'Report/geral.html', context)

@login_required
@allowed_users(allowed_roles=['postu', 'suku', 'ald', 'admin', 'staff'])
def report_posisaun(request):
    group = request.user.groups.all()[0].name
    user_mem = c_user_mem(request.user)
    loc = None
    if user_mem:
        loc = LocationTL.objects.select_related('village', 'aldeia').filter(membro=user_mem).first()

    qs, tipe = get_role_queryset(group, loc)

    from django.db.models import Count
    rows = (
        qs.values('membro__membroposition__position__name')
          .annotate(total=Count('id'))
          .order_by('-total')
    )
    data = [
        {
            'posisaun': r['membro__membroposition__position__name'] or '—',
            'total': r['total']
        }
        for r in rows
    ]

    context = {
        'title': 'Relatoriu Posisaun',
        'legend': 'Relatoriu Per Posisaun',
        'group': group,
        'data': data,
        'total': qs.count(),
    }
    return render(request, 'Report/posisaun.html', context)

@login_required
@allowed_users(allowed_roles=['postu', 'suku', 'ald', 'admin', 'staff'])
def report_status(request):
    group = request.user.groups.all()[0].name
    user_mem = c_user_mem(request.user)
    loc = None
    if user_mem:
        loc = LocationTL.objects.select_related('village', 'aldeia').filter(membro=user_mem).first()

    qs, tipe = get_role_queryset(group, loc)

    from django.db.models import Count
    rows = (
        qs.values('membro__status__name')
          .annotate(total=Count('id'))
          .order_by('-total')
    )
    data = [
        {'status': r['membro__status__name'] or '—', 'total': r['total']}
        for r in rows
    ]

    context = {
        'title': 'Relatoriu Status',
        'legend': 'Relatoriu Per Status',
        'group': group,
        'data': data,
        'total': qs.count(),
    }
    return render(request, 'Report/status.html', context)